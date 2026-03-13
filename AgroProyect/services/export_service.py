"""
Módulo de Exportación de Datos
Genera reportes en Excel, PDF con gráficos y documentos profesionales
Con marca de agua translúcida del logo Agro-Master
"""

from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import os
from datetime import datetime
import matplotlib
matplotlib.use('Agg') # Forzar backend no interactivo antes de importar pyplot
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from io import BytesIO
import base64
from PIL import Image

# Ruta al favicon/logo para marca de agua
LOGO_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'static', 'img', 'favicon.png')


class PDFConMarcaAgua(FPDF):
    """
    MOTOR DE RENDERIZADO PDF (CUSTOM):
    Extiende la funcionalidad de FPDF para inyectar una capa de identidad visual proactiva.
    
    Características:
    - Marca de Agua Dinámica: Inserta el logo corporativo con opacidad controlada (15%-25%) 
      en el centro de cada página de forma automática mediante el hook del 'footer'.
    - Gestión de Transparencias: Utiliza Pillow (PIL) para pre-procesar el canal alfa 
      de la imagen antes de insertarla en el canvas del PDF.
    """
    
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.logo_path = LOGO_PATH
        self.marca_agua_opacidad = 0.15  # 15% de opacidad para la marca de agua (más visible)
        self._watermark_image = None
        self._watermark_bytes = None  # Guardar bytes para reutilizar en cada página
        self._preparar_marca_agua()
    
    def _preparar_marca_agua(self):
        """Prepara la imagen de marca de agua con transparencia"""
        try:
            if os.path.exists(self.logo_path):
                # Abrir imagen y asegurar que tenga canal alfa
                img = Image.open(self.logo_path).convert('RGBA')
                
                # Ajustar la opacidad (0.25 para que sea claramente visible sobre el contenido)
                alpha_factor = 0.25 
                
                # Procesar pixeles para ajustar transparencia
                r, g, b, a = img.split()
                a = a.point(lambda p: int(p * alpha_factor))
                self._watermark_pil = Image.merge('RGBA', (r, g, b, a))
                self._watermark_image = True
            else:
                # Buscar en rutas alternativas si falla la principal
                alt_path = os.path.join('static', 'img', 'favicon.png')
                if os.path.exists(alt_path):
                    self.logo_path = alt_path
                    self._preparar_marca_agua()
        except Exception as e:
            print(f"Error preparando marca de agua: {e}")
            self._watermark_image = False
    
    def header(self):
        """Encabezado normal (sin marca de agua aquí para que no quede debajo)"""
        pass
    
    def footer(self):
        """Pie de página profesional y marca de agua (dibujada al final para estar encima)"""
        # --- DIBUJAR MARCA DE AGUA (SOBRE TODO EL CONTENIDO) ---
        if self._watermark_image and hasattr(self, '_watermark_pil'):
            try:
                # Guardar estado de posición actual
                curr_x, curr_y = self.get_x(), self.get_y()
                
                page_width, page_height = self.w, self.h
                watermark_size = min(page_width, page_height) * 0.55
                x = (page_width - watermark_size) / 2
                y = (page_height - watermark_size) / 2
                
                self.image(self._watermark_pil, x=x, y=y, w=watermark_size, h=watermark_size)
                
                # Restaurar posición para el footer normal
                self.set_xy(curr_x, curr_y)
            except Exception as e:
                print(f"Error agregando marca de agua en footer: {e}")

        # --- PIE DE PÁGINA NORMAL ---
        self.set_y(-15)
        self.set_font('Arial', 'I', 8)
        self.set_text_color(128, 128, 128)
        self.cell(0, 10, f'Agro-Master © {datetime.now().year} - Página {self.page_no()}/{{nb}}', 0, 0, 'C')
        self.set_text_color(0, 0, 0)

class ExportadorExcel:
    """
    GENERADOR DE LIBROS DE TRABAJO (XLSX):
    Construye reportes tabulares altamente estructurados utilizando openpyxl.
    
    Diseño Visual:
    - Aplicación de estilos CSS-like (Font, Fill, Alignment) sobre celdas de Excel.
    - Encabezados estáticos fusionados para branding.
    - Generación de hojas múltiples (Inventario + Estadísticas).
    """
    
    def __init__(self, titulo="Reporte Ganado"):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Inventario Ganado"
        self.titulo = titulo
        self.crear_encabezado()
    
    def crear_encabezado(self):
        """Crea el encabezado profesional del Excel"""
        # Fusionar celdas para título
        self.ws.merge_cells('A1:H1')
        titulo_cell = self.ws['A1']
        titulo_cell.value = "AGRO-MASTER"
        titulo_cell.font = Font(bold=True, size=16, color="FFFFFF")
        titulo_cell.fill = PatternFill(start_color="1b4332", end_color="1b4332", fill_type="solid")
        titulo_cell.alignment = Alignment(horizontal="center", vertical="center")
        self.ws.row_dimensions[1].height = 25
        
        # Subtítulo
        self.ws.merge_cells('A2:H2')
        subtitulo_cell = self.ws['A2']
        subtitulo_cell.value = "Sistema de Gestión Ganadera - " + self.titulo
        subtitulo_cell.font = Font(size=11, color="FFFFFF")
        subtitulo_cell.fill = PatternFill(start_color="2d6a4f", end_color="2d6a4f", fill_type="solid")
        subtitulo_cell.alignment = Alignment(horizontal="center", vertical="center")
        self.ws.row_dimensions[2].height = 20
        
        # Fecha de generación
        self.ws.merge_cells('A3:H3')
        fecha_cell = self.ws['A3']
        fecha_cell.value = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        fecha_cell.font = Font(size=9, italic=True)
        fecha_cell.alignment = Alignment(horizontal="right")
        
        # Encabezados de tabla
        headers = ['ID', 'Código', 'Especie', 'Raza', 'Sexo', 'Edad (meses)', 'Peso (kg)', 'Estado', 'Padre', 'Madre']
        header_fill = PatternFill(start_color="ffca3a", end_color="ffca3a", fill_type="solid")
        header_font = Font(bold=True, color="000000")
        
        for col_num, header in enumerate(headers, 1):
            cell = self.ws.cell(row=5, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        self.ws.row_dimensions[5].height = 18
    
    def agregar_datos(self, animales):
        """Agrega los datos de los animales al Excel"""
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for idx, animal in enumerate(animales, start=6):
            self.ws.cell(row=idx, column=1).value = animal.get('id', 'N/A')
            self.ws.cell(row=idx, column=2).value = animal.get('codigo_identificacion', 'N/A')
            self.ws.cell(row=idx, column=3).value = animal.get('especie', 'N/A')
            self.ws.cell(row=idx, column=4).value = animal.get('raza', 'N/A')
            self.ws.cell(row=idx, column=5).value = animal.get('sexo', 'N/A')
            self.ws.cell(row=idx, column=6).value = animal.get('edad', 0)
            self.ws.cell(row=idx, column=7).value = animal.get('peso', 0.0)
            self.ws.cell(row=idx, column=8).value = animal.get('estado', 'N/A')
            self.ws.cell(row=idx, column=9).value = animal.get('padre_id', '---')
            self.ws.cell(row=idx, column=10).value = animal.get('madre_id', '---')
            
            # Aplicar bordes y alineación
            for col in range(1, 11):
                cell = self.ws.cell(row=idx, column=col)
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Ajustar ancho de columnas
        self.ws.column_dimensions['A'].width = 8
        self.ws.column_dimensions['B'].width = 15
        self.ws.column_dimensions['C'].width = 12
        self.ws.column_dimensions['D'].width = 18
        self.ws.column_dimensions['E'].width = 10
        self.ws.column_dimensions['F'].width = 13
        self.ws.column_dimensions['G'].width = 12
        self.ws.column_dimensions['H'].width = 15
        self.ws.column_dimensions['I'].width = 10
        self.ws.column_dimensions['J'].width = 10
    
    def agregar_estadisticas(self, total, peso_promedio, alertas, criticos):
        """Agrega una hoja de estadísticas"""
        ws_stats = self.wb.create_sheet("Estadísticas")
        
        # Título
        ws_stats.merge_cells('A1:B1')
        titulo = ws_stats['A1']
        titulo.value = "ESTADÍSTICAS"
        titulo.font = Font(bold=True, size=14, color="FFFFFF")
        titulo.fill = PatternFill(start_color="1b4332", end_color="1b4332", fill_type="solid")
        titulo.alignment = Alignment(horizontal="center")
        
        # Datos
        datos = [
            ["Total de Animales", total],
            ["Peso Promedio (kg)", f"{peso_promedio:.2f}"],
            ["Alertas (bajo peso)", alertas],
            ["Críticos", criticos]
        ]
        
        for idx, fila in enumerate(datos, start=3):
            ws_stats.cell(row=idx, column=1).value = fila[0]
            ws_stats.cell(row=idx, column=1).font = Font(bold=True)
            ws_stats.cell(row=idx, column=2).value = fila[1]
            ws_stats.cell(row=idx, column=2).font = Font(size=12, bold=True)
        
        ws_stats.column_dimensions['A'].width = 25
        ws_stats.column_dimensions['B'].width = 15
    
    def obtener_bytes(self):
        """Retorna el contenido del Excel como bytes"""
        output = io.BytesIO()
        self.wb.save(output)
        output.seek(0)
        return output

class ExportadorPDF:
    """Genera reportes profesionales en PDF con gráficos y marca de agua"""
    
    def __init__(self, titulo="Reporte de Ganado", incluir_grafica=True, descripcion=None):
        self.pdf = PDFConMarcaAgua()
        self.pdf.alias_nb_pages()  # Para el total de páginas en el footer
        # Agregar fuente Unicode para soportar caracteres especiales
        try:
            self.pdf.add_font('DejaVu', '', 'DejaVuSans.ttf', uni=True)
            self.font_unicode = 'DejaVu'
        except:
            self.font_unicode = 'Arial'  # Fallback
        self.pdf.add_page()
        self.titulo = titulo
        self.descripcion = descripcion or "Sistema Integral de Gestión de Producción Agropecuaria Avanzada."
        self.incluir_grafica = incluir_grafica
        self.crear_encabezado()
    
    def crear_encabezado(self):
        """Crea el encabezado profesional (sin logo, solo marca de agua de fondo)"""
        # Título principal
        self.pdf.set_fill_color(27, 67, 50)  # Verde oscuro corporativo
        self.pdf.set_text_color(255, 255, 255)
        self.pdf.set_font("Arial", 'B', 18)
        self.pdf.cell(0, 12, txt="AGRO-MASTER", ln=True, align='C', fill=True)
        
        # Subtítulo
        self.pdf.set_fill_color(45, 106, 79)  # Verde medio
        self.pdf.set_font("Arial", '', 11)
        self.pdf.cell(0, 8, txt="Sistema de Gestion Ganadera - " + self.titulo, ln=True, align='C', fill=True)
        
        # Descripción del documento
        if self.descripcion:
            self.pdf.set_fill_color(240, 253, 244)  # Verde muy claro
            self.pdf.set_text_color(34, 85, 60)
            self.pdf.set_font("Arial", 'I', 9)
            self.pdf.multi_cell(0, 5, txt=self.descripcion, align='C', fill=True)
            self.pdf.set_text_color(0, 0, 0)
        
        # Línea separadora decorativa
        self.pdf.set_draw_color(27, 67, 50)
        self.pdf.set_line_width(0.5)
        self.pdf.line(10, self.pdf.get_y() + 2, 200, self.pdf.get_y() + 2)
        
        # Fecha y hora de generación
        self.pdf.ln(5)
        self.pdf.set_fill_color(255, 255, 255)
        self.pdf.set_text_color(100, 100, 100)
        self.pdf.set_font("Arial", 'I', 8)
        self.pdf.cell(0, 5, txt=f"Documento generado el: {datetime.now().strftime('%d/%m/%Y a las %H:%M:%S')}", ln=True, align='R')
        self.pdf.set_text_color(0, 0, 0)
        
        self.pdf.ln(3)
    
    def agregar_tabla(self, headers, datos):
        """Agrega una tabla al PDF"""
        col_width = 190 / len(headers)
        
        # Encabezados
        self.pdf.set_fill_color(255, 202, 58)  # Amarillo
        self.pdf.set_text_color(0, 0, 0)
        self.pdf.set_font("Arial", 'B', 9)
        
        for header in headers:
            self.pdf.cell(col_width, 7, txt=header, border=1, align='C', fill=True)
        self.pdf.ln()
        
        # Datos
        self.pdf.set_font("Arial", '', 8)
        self.pdf.set_text_color(0, 0, 0)
        
        for fila in datos:
            for valor in fila:
                self.pdf.cell(col_width, 6, txt=str(valor), border=1, align='C')
            self.pdf.ln()
    
    def agregar_grafica(self, etiquetas, valores, titulo_grafica):
        """Agrega una gráfica al PDF"""
        try:
            # Crear figura pequeña
            fig, ax = plt.subplots(figsize=(6, 3))
            colors = ['#2d6a4f', '#1b4332', '#ffca3a', '#ff9d00', '#d62828']
            ax.bar(etiquetas, valores, color=colors[:len(etiquetas)])
            ax.set_title(titulo_grafica, fontsize=12, fontweight='bold')
            ax.set_ylabel('Cantidad', fontsize=10)
            ax.grid(axis='y', alpha=0.3)
            
            # Guardar en bytes
            buf = BytesIO()
            plt.tight_layout()
            fig.savefig(buf, format='png', dpi=100, bbox_inches='tight')
            buf.seek(0)
            plt.close(fig)
            
            # Insertar en PDF
            img_path = buf
            self.pdf.image(img_path, x=20, w=170)
            self.pdf.ln(5)
            
        except Exception as e:
            print(f"Error al agregar gráfica: {e}")
    
    def agregar_estadisticas(self, stats_dict):
        """Agrega sección de estadísticas"""
        self.pdf.set_font("Arial", 'B', 12)
        self.pdf.cell(0, 8, txt="Estadisticas Generales", ln=True, border=0)
        self.pdf.ln(2)
        
        self.pdf.set_font("Arial", '', 10)
        for clave, valor in stats_dict.items():
            self.pdf.cell(0, 6, txt=f"- {clave}: {valor}", ln=True)
        
        self.pdf.ln(5)
    
    def agregar_datos(self, animales):
        """Agrega los datos de los animales en formato tabla"""
        headers = ['ID', 'Código', 'Especie', 'Edad', 'Peso', 'Estado']
        datos = []
        for animal in animales:
            fila = [
                animal.get('id', 'N/A'),
                animal.get('codigo_identificacion', 'N/A'),
                animal.get('especie', 'N/A'),
                f"{animal.get('edad', 0)}m",
                f"{animal.get('peso', 0)}kg",
                animal.get('estado', 'N/A')
            ]
            datos.append(fila)
        
        self.agregar_tabla(headers, datos)
    
    def obtener_bytes(self):
        """Retorna el PDF como bytes"""
        return io.BytesIO(self.pdf.output(dest='S'))

def generar_grafica_especie(animales):
    """Genera una gráfica de distribución por especie"""
    from collections import Counter
    
    especies = [a['especie'] for a in animales]
    conteo = Counter(especies)
    
    fig, ax = plt.subplots(figsize=(8, 5))
    colors = ['#2d6a4f', '#1b4332', '#ffca3a', '#ff9d00', '#d62828']
    ax.bar(conteo.keys(), conteo.values(), color=colors[:len(conteo)])
    ax.set_title('Distribución de Ganado por Especie', fontsize=14, fontweight='bold')
    ax.set_ylabel('Cantidad', fontsize=11)
    ax.set_xlabel('Especie', fontsize=11)
    ax.grid(axis='y', alpha=0.3)
    
    buf = BytesIO()
    plt.tight_layout()
    fig.savefig(buf, format='png', dpi=100, bbox_inches='tight')
    buf.seek(0)
    plt.close(fig)
    
    return buf

def generar_grafica_pesos(animales):
    """Genera una gráfica de distribución de pesos"""
    import numpy as np
    
    pesos = [a['peso'] for a in animales]
    
    fig, ax = plt.subplots(figsize=(8, 5))
    ax.hist(pesos, bins=10, color='#2d6a4f', edgecolor='#1b4332', alpha=0.7)
    ax.set_title('Distribución de Pesos', fontsize=14, fontweight='bold')
    ax.set_xlabel('Peso (kg)', fontsize=11)
    ax.set_ylabel('Cantidad de Animales', fontsize=11)
    ax.grid(axis='y', alpha=0.3)
    
    buf = BytesIO()
    plt.tight_layout()
    fig.savefig(buf, format='png', dpi=100, bbox_inches='tight')
    buf.seek(0)
    plt.close(fig)
    
    return buf
